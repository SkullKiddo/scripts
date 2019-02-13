from openpyxl import Workbook
from openpyxl.compat import range
from openpyxl.utils import get_column_letter
from openpyxl import load_workbook
from openpyxl.compat import range
import os
import math
entrada = os.path.join(os.getcwd(),"nombre entrada")
sortida = os.path.join(os.getcwd(),"nombre salida")
nombre = 0
fecha = 1
importe = 4
wb = Workbook()
entradaWB = load_workbook(entrada)
entrada = entradaWB['nombre pestaña entrada']
resultado = wb.active
resultado.title = "Movimientos agrupados"



def letra(n):
    if n > 25:
        l1,l2 = math.floor(n/26), n%26
        return chr(l1+64)+chr(l2+65)
    return chr(n+65)


def escribe(dato,x,y):
    debug = dato
    pos = letra(x)+str(y)
    resultado[pos] = dato


def equivalente(name):
    if name[:3] == "GNS" and name[3] >= '0' and name[3] <= '9':
        return "GAS NATURAL"
    return name


colmap = []
y = 2
x = 0
fechaAnterior = entrada['B1'].value
for movimiento in entrada:
    name = movimiento[nombre].value
    name = equivalente(name)
    if not name in colmap:
        colmap.append(name)
        escribe(name, colmap.index(name) +1, 1)
    x = colmap.index(name) +1 #dejo espacio para la fecha
    ''' ESTO SIRVE PARA QUE LOS MOVIMIENTOS DEL MISMO DIA QUEDEN EN LA MISMA COLUMNA, habria que quitar el otro y++
    if movimiento[fecha].value != fechaAnterior:
        fechaAnterior = movimiento[fecha].value
        y += 1
    '''
    y += 1
    escribe(movimiento[fecha].value, 0, y)
    escribe(movimiento[importe].value, x, y)
wb.save(filename="Movimientos_cuenta.xlsx")    
